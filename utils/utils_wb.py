import os

import datetime
import numpy as np
import pandas as pd

# Needed for excel tables
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM
from openpyxl.styles import Alignment, borders, Font

from typing import Any, Set, Union, List, Dict

font_bold = Font(bold=True, color='00000000')
alignment_center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
border_medium_t_b_l_r = Border(
    left=Side(border_style=BORDER_MEDIUM, color='00000000'),
    right=Side(border_style=BORDER_MEDIUM, color='00000000'),
    top=Side(border_style=BORDER_MEDIUM, color='00000000'),
    bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
)
thin_border = borders.Border(
    left=borders.Side(border_style=borders.BORDER_THIN, color='00000000'),
    right=borders.Side(border_style=borders.BORDER_THIN, color='00000000'),
    top=borders.Side(border_style=borders.BORDER_THIN, color='00000000'),
    bottom=borders.Side(border_style=borders.BORDER_THIN, color='00000000')
)
thin_border_gray = borders.Border(
    left=borders.Side(border_style=borders.BORDER_THIN, color='FFC5C5C5'),
    right=borders.Side(border_style=borders.BORDER_THIN, color='FFC5C5C5'),
    top=borders.Side(border_style=borders.BORDER_THIN, color='FFC5C5C5'),
    bottom=borders.Side(border_style=borders.BORDER_THIN, color='FFC5C5C5')
)
thin_border_t = Border(
    top=Side(border_style=BORDER_THIN, color='00000000')
)


def convert_sheet_to_df(wb, sheet_name):
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)
    df = pd.DataFrame(df.values[1:], columns=df.values[0])
    return df


def get_column_widths_file(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    widths = []
    for i in range(1, ws.max_column + 1):
        widths.append(ws.column_dimensions[get_column_letter(i)].width)
    return widths


def get_column_widths_wb(wb, sheet_name):
    ws = wb[sheet_name]
    widths = []
    for i in range(1, ws.max_column + 1):
        widths.append(ws.column_dimensions[get_column_letter(i)].width)
    return widths


def my_ceil(arr, prec: int=0) -> Union[np.ndarray, Any]:
    return np.true_divide(np.ceil(arr * 10**prec), 10**prec)


def get_all_column_widths_file(file_path):
    d_width = {}
    d_column_width = {}
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        widths = []
        column_widths = []
        width_prev = 13.0
        for i in range(1, ws.max_column + 1):
            width = ws.column_dimensions[get_column_letter(i)].width

            if width == 13.0:
                width = width_prev
            else:
                width_prev = my_ceil(width, prec=2)

            widths.append(width)
            column = ws.cell(row=1, column=i).value
            column_widths.append((column, width))
        d_width[sheet_name] = widths
        d_column_width[sheet_name] = column_widths
        print("sheet_name: {}, widths: {}, column_widths: {}".format(sheet_name, widths, column_widths))
    return d_width, d_column_width


def create_new_sheet(wb, sheet_name, list_build, column_widths, wrap_first_row=False, first_row_height=15, freeze_row=1,
                     freeze_column=1, columns_idx_float: Set = set(), columns_idx_percent: Set = set()):
    print(f"Creating a new sheet with the name '{sheet_name}'")
    
    ws = wb.create_sheet(sheet_name)
    cl = ws.cell

    for i, col_name in enumerate(list_build[0], 1):
        c = cl(row=1, column=i)
        c.value = col_name
    for j, row_vals in enumerate(list_build[1:], 2):
        for i, val in enumerate(row_vals, 1):
            c = cl(row=j, column=i)
            if isinstance(val, str) and '%' in val and '.' in val:
                c.value = float(val.strip('%')) / 100
                c.number_format = '0.0000%'
            elif i in columns_idx_percent:
                c.value = val
                c.number_format = '0.0000%'
            elif isinstance(val, float) or i in columns_idx_float:
                c.value = val
                c.number_format = '0.00'
            else:
                c.value = val

    if wrap_first_row:
        for i in range(1, len(list_build[0]) + 1):
            c = cl(row=1, column=i)
            c.border = thin_border
            c.alignment = alignment_center_wrap
            c.font = font_bold
    ws.row_dimensions[1].height = first_row_height

    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width + 0.4

    ws.freeze_panes = "{}{}".format(get_column_letter(freeze_column + 1), freeze_row + 1)
    ws.sheet_view.zoomScale = 80

    ws.auto_filter.ref = ws.dimensions


def get_whole_xlsx_content(file_path: str) -> Dict[str, List[List[Any]]]:
    assert os.path.exists(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    d_l_content = {}
    for sheetname in wb.sheetnames:
        l_content = list(wb[sheetname].values)
        d_l_content[sheetname] = l_content

    wb.close()

    return d_l_content


def read_all_df_from_xlsx(file_path: str) -> Dict[str, List[List[Any]]]:
    assert os.path.exists(file_path)
    d_l_content = get_whole_xlsx_content(file_path)

    d_df: Dict[str, pd.DataFrame] = {}
    for sheetname, l_content in d_l_content.items():
        d_df[sheetname] = pd.DataFrame(data=l_content[1:], columns=l_content[0], dtype=object)

    return d_df


def convert_xlsx_days_to_datetime(days: int) -> datetime.datetime:
    dt_delta = datetime.timedelta(days=days)
    dt = datetime.datetime.strptime('1899/12/30','%Y/%m/%d') + dt_delta

    return dt
