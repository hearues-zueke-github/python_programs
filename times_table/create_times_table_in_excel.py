#! /usr/bin/python3.14

import openpyxl
import qrcode

import numpy as np

from copy import copy

from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils.units import pixels_to_EMU

MM_PER_INCH = 25.4
INCH_PER_POINT = 1. / 72

POINT_PER_PIXEL = 12.75 / 17.
CHARACTER_PER_PIXEL = 8.43 / 64

func_row_height_point_in_mm = lambda x_mm: (
	x_mm / MM_PER_INCH / INCH_PER_POINT
)
# func_column_width_chars_in_mm = lambda x_mm: (
# 	(
# 		x_mm
# 		/ MM_PER_INCH
# 		/ INCH_PER_POINT
# 		/ POINT_PER_PIXEL
# 		- 5
# 	)
# 	* CHARACTER_PER_PIXEL
# 	* 0.80256821
# )
func_column_width_chars_in_mm = lambda x_mm: (
	# 0.10018905 + x_mm * 0.377928949
	0.003779289 + x_mm * 0.377928949
)

# char_3 = 10.05 mm
# char 8 = 23.28 mm
# char 15 = 41.80 mm


if __name__ == '__main__':
	print('Hello World!')

	arr_col = np.arange(1, 11)
	np.random.shuffle(arr_col)

	arr_row = np.arange(1, 11)
	np.random.shuffle(arr_row)

	arr_times_table = arr_row.reshape((10, -1)) * arr_col.reshape((-1, 10))

	wb = openpyxl.Workbook()

	del wb['Sheet']
	ws = wb.create_sheet('Times Table 1')

	ws.page_setup.paperSize = ws.PAPERSIZE_A4

	# Set margins (values are in inches)
	ws.page_margins.left = 0.5
	ws.page_margins.right = 0.5
	ws.page_margins.top = 0.5
	ws.page_margins.bottom = 0.5
	ws.page_margins.header = 0.05
	ws.page_margins.footer = 0.05

	ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

	# ws.sheet_view.view = "pageBreakPreview"

	font_size_symbol = 28
	font_size_border = 16
	font_size_inner_cells = 12

	ws.sheet_properties.pageSetUpPr = PageSetupProperties(autoPageBreaks=True)

	row_heights_mm = [10 for i in range(1, 12)]
	row_heights = [func_row_height_point_in_mm(x_mm=v) for v in row_heights_mm]
	# column_widths = [func_column_width_chars_in_mm(x_mm=25)]*4
	column_widths_mm = [10 for i in range(1, 12)]
	column_widths = [func_column_width_chars_in_mm(x_mm=v) for v in column_widths_mm]
	# column_widths = [v for v in column_widths_mm]
	
	for i, column_width in enumerate(column_widths, 1):
		ws.column_dimensions[get_column_letter(i)].width = column_width

	for i, row_height in enumerate(row_heights, 1):
		ws.row_dimensions[i].height = row_height

	thin = Side(border_style="thin", color="000000")
	thick = Side(border_style="thick", color="000000")
	border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
	border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)

	c = ws.cell(row=1, column=1)
	c.value = 'x'
	c.alignment = Alignment(horizontal='center', vertical='center')
	c.border = border_thin

	for i, v in enumerate(arr_row, 2):
		c = ws.cell(row=i, column=1)
		c.value = v
		c.alignment = Alignment(horizontal='center', vertical='center')
		c.border = border_thin

	for i, v in enumerate(arr_col, 2):
		c = ws.cell(row=1, column=i)
		c.value = v
		c.alignment = Alignment(horizontal='center', vertical='center')
		c.border = border_thin

	for i, arr_row_times in enumerate(arr_times_table, 2):
		for j, v in enumerate(arr_row_times, 2):
			c = ws.cell(row=i, column=j)
			c.value = v
			c.alignment = Alignment(horizontal='center', vertical='center')
			c.border = border_thin

	for i in range(1, 12):
		c = ws.cell(row=1, column=i)
		border = copy(c.border)
		border.top = thick
		border.bottom = thick
		c.border = border

		c = ws.cell(row=11, column=i)
		border = copy(c.border)
		border.bottom = thick
		c.border = border

	for i in range(1, 12):
		c = ws.cell(row=i, column=1)
		border = copy(c.border)
		border.left = thick
		border.right = thick
		c.border = border

		c = ws.cell(row=i, column=11)
		border = copy(c.border)
		border.right = thick
		c.border = border

	c = ws.cell(row=1, column=1)
	font = copy(c.font)
	font.size = font_size_symbol
	font.bold = True
	c.font = font

	for i in range(1, 11):
		c = ws.cell(row=1+i, column=1)
		font = copy(c.font)
		font.size = font_size_border
		font.bold = True
		c.font = font

	for i in range(1, 11):
		c = ws.cell(row=1, column=1+i)
		font = copy(c.font)
		font.size = font_size_border
		font.bold = True
		c.font = font

	for j in range(1, 11):
		for i in range(1, 11):
			c = ws.cell(row=1+j, column=1+i)
			font = copy(c.font)
			font.size = font_size_inner_cells
			font.bold = False
			c.font = font


	# Empty field for writting
	row_offset = 13
	# for i, column_width in enumerate(column_widths, 1):
	# 	ws.column_dimensions[get_column_letter(i)].width = column_width

	for i, row_height in enumerate(row_heights, row_offset+1):
		ws.row_dimensions[i].height = row_height

	thin = Side(border_style="thin", color="000000")
	thick = Side(border_style="thick", color="000000")
	border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
	border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)

	c = ws.cell(row=row_offset+1, column=1)
	c.value = 'x'
	c.alignment = Alignment(horizontal='center', vertical='center')
	c.border = border_thin

	for i, v in enumerate(arr_row, 2):
		c = ws.cell(row=row_offset+i, column=1)
		c.value = v
		c.alignment = Alignment(horizontal='center', vertical='center')
		c.border = border_thin

	for i, v in enumerate(arr_col, 2):
		c = ws.cell(row=row_offset+1, column=i)
		c.value = v
		c.alignment = Alignment(horizontal='center', vertical='center')
		c.border = border_thin

	for i, arr_row_times in enumerate(arr_times_table, 2):
		for j, v in enumerate(arr_row_times, 2):
			c = ws.cell(row=row_offset+i, column=j)
			# c.value = v
			c.alignment = Alignment(horizontal='center', vertical='center')
			c.border = border_thin

	for i in range(1, 12):
		c = ws.cell(row=row_offset+1, column=i)
		border = copy(c.border)
		border.top = thick
		border.bottom = thick
		c.border = border

		c = ws.cell(row=row_offset+11, column=i)
		border = copy(c.border)
		border.bottom = thick
		c.border = border

	for i in range(1, 12):
		c = ws.cell(row=row_offset+i, column=1)
		border = copy(c.border)
		border.left = thick
		border.right = thick
		c.border = border

		c = ws.cell(row=row_offset+i, column=11)
		border = copy(c.border)
		border.right = thick
		c.border = border

	c = ws.cell(row=row_offset+1, column=1)
	font = copy(c.font)
	font.size = font_size_symbol
	font.bold = True
	c.font = font

	for i in range(1, 11):
		c = ws.cell(row=row_offset+1+i, column=1)
		font = copy(c.font)
		font.size = font_size_border
		font.bold = True
		c.font = font

	for i in range(1, 11):
		c = ws.cell(row=row_offset+1, column=1+i)
		font = copy(c.font)
		font.size = font_size_border
		font.bold = True
		c.font = font


	qr = qrcode.QRCode(
		version=10,
		error_correction=qrcode.constants.ERROR_CORRECT_H,
		box_size=10,
		border=4,
	)
	data = (
		'numbers_in_1st_col = [' + ', '.join(str(v) for v in arr_row) + ']; ' +
		'numbers_in_1st_row = [' + ', '.join(str(v) for v in arr_col) + ']'
	)
	print(f'len(data): {len(data)}')
	qr.add_data(data)
	qr.make(fit=True)

	img = qr.make_image(fill_color="black", back_color="white")
	img.save('qr_logo.png')
	
	# img2 = Image("logo.png")
	img2 = Image("qr_logo.png")

	anchor = AbsoluteAnchor(
		pos=XDRPoint2D(pixels_to_EMU(450), pixels_to_EMU(140)),
		ext=XDRPositiveSize2D(pixels_to_EMU(img2.width)/3., pixels_to_EMU(img2.height)/3.)
	)

	ws.add_image(img2, anchor)

	wb.save('times_table.xlsx')
