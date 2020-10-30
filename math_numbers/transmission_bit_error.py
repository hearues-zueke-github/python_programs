#! /usr/bin/python3.8

# -*- coding: utf-8 -*-

# Some other needed imports
import datetime
import dill
import gzip
import os
import pdb
import re
import sys
import traceback

import numpy as np
import pandas as pd
from pprint import pprint

from copy import deepcopy, copy
from dotmap import DotMap
from functools import reduce
from memory_tempfile import MemoryTempfile
from shutil import copyfile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

from openpyxl.styles.borders import Border, Side, BORDER_THICK

thin_border = Border(
    # left=Side(border_style=BORDER_THIN, color='00000000'),
    # right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THICK, color='00000000'),
    # bottom=Side(border_style=BORDER_THIN, color='00000000')
)
# rkbook()
# ws = wb.get_active_sheet()
# # property cell.border should be used instead of cell.style.border
# ws.cell(row=3, column=2).border = thin_border
# wb.save('b
# ws.cell(row=3, column=2).border = thin_border

# thin_border = Border(left=Side(style='thin'), 
#                      right=Side(style='thin'), 
#                      top=Side(style='thin'), 
#                      bottom=Side(style='thin'))

# wb = Workbook()
# ws = wb.get_active_sheet()
# # property cell.border should be used instead of cell.style.border
# ws.cell(row=3, column=2).border = thin_border
# wb.save('border_test.xlsx')

from typing import List, Tuple

PATH_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))+"/"
HOME_DIR = os.path.expanduser("~")+"/"
TEMP_DIR = MemoryTempfile().gettempdir()+"/"


def mkdirs(path):
    if not os.path.exists(path):
        os.makedirs(path)
# from openpyxl.styles.borders import Border, Side
# from openpyxl import Workbook

# thin_border = Border(left=Side(style='thin'), 
#                      right=Side(style='thin'), 
#                      top=Side(style='thin'), 
#                      bottom=Side(style='thin'))

# wb = Workbook()
# ws = wb.get_active_sheet()
# # property cell.border should be used instead of cell.style.border
# ws.cell(row=3, column=2).border = thin_border
# wb.save('border_test.xlsx')


OBJS_DIR_PATH = PATH_ROOT_DIR+'objs/'
mkdirs(OBJS_DIR_PATH)

XLSX_TEMP_DIR = os.path.join(TEMP_DIR, 'xlsx_files')+'/'
mkdirs(XLSX_TEMP_DIR)


def create_orig_fail_msg(n: int) -> Tuple[List[int], List[int], pd.DataFrame]:
    arr_orig = np.random.randint(0, 2, (n, ), dtype=np.uint8)
    # arr_fails_idx = np.where(np.random.randint(0, 20, (n, )) == 0)[0]
    arr_fails_idx = np.where(np.random.randint(0, 100, (n, )) == 0)[0]
    arr_fail_type = np.random.randint(0, 3, (arr_fails_idx.shape[0], ))

    arr_fails_idx_type_0 = arr_fails_idx[arr_fail_type == 0]
    arr_fails_idx_type_1 = arr_fails_idx[arr_fail_type == 1]
    arr_fails_idx_type_2 = arr_fails_idx[arr_fail_type == 2]

    # type 0... bit flip
    # type 1... bit delete
    # type 2... bit insert

    df = pd.DataFrame(data={'bits': arr_orig}, columns=['bits'])
    df.index = np.arange(0, df.shape[0]*2, 2)

    arr_index_orig = df.index.values

    # flip the bits!
    df.loc[arr_index_orig[arr_fails_idx_type_0], 'bits'] = df.loc[arr_index_orig[arr_fails_idx_type_0], 'bits'].values ^ 1
    df['flip_bit'] = 0
    df.loc[arr_index_orig[arr_fails_idx_type_0], 'flip_bit'] = 1

    # delete the bits!
    df.drop(index=arr_index_orig[arr_fails_idx_type_1], inplace=True)
    df['delete_bit_prev'] = 0
    arr_available_idx = df.index.values[np.isin(df.index.values, arr_index_orig[arr_fails_idx_type_1]+2)]
    df.loc[arr_available_idx, 'delete_bit_prev'] = 1

    # insert the bits!
    arr_fails_insert_pos = np.random.randint(0, 2, (arr_fails_idx_type_2.shape[0], ))
    arr_index_insert_bit_prev = arr_fails_idx_type_2[arr_fails_insert_pos == 0]
    arr_index_insert_bit_next = arr_fails_idx_type_2[arr_fails_insert_pos == 1]

    arr_insert_bit_prev = np.random.randint(0, 2, (arr_index_insert_bit_prev.shape[0], ), dtype=np.uint8)
    arr_insert_bit_next = np.random.randint(0, 2, (arr_index_insert_bit_next.shape[0], ), dtype=np.uint8)

    for idx in arr_index_orig[arr_fails_idx_type_2] + 1:
        df.loc[int(idx)] = [None]*df.shape[1]
    df.sort_index(inplace=True)

    df['insert_bit'] = 0

    arr_idx_prev = arr_index_orig[arr_index_insert_bit_prev]
    df.loc[arr_idx_prev + 1, 'bits'] = df.loc[arr_idx_prev, 'bits'].values
    df.loc[arr_idx_prev, 'bits'] = arr_insert_bit_prev

    arr_idx_next = arr_index_orig[arr_index_insert_bit_next]
    df.loc[arr_idx_next + 1, 'bits'] = arr_insert_bit_next

    df.loc[arr_idx_prev, 'insert_bit'] = 1
    df.loc[arr_idx_next + 1, 'insert_bit'] = 1

    arr_fail = df['bits'].values

    return arr_orig, arr_fail, df


if __name__ == '__main__':
    print("Hello World!")

    n = 1000

    arr_orig, arr_fail, df_orig = create_orig_fail_msg(n=n)

    l_orig = arr_orig.tolist()
    l_fail = arr_fail.tolist()

    len_l_orig = len(l_orig)
    len_l_fail = len(l_fail)
    max_n = max(len_l_orig, len_l_fail)

    if len_l_orig < max_n:
        l_orig += [None] * (max_n - len_l_orig)
    if len_l_fail < max_n:
        l_fail += [None] * (max_n - len_l_fail)

    df = pd.DataFrame(
        data={'orig': l_orig, 'fail': l_fail},
        columns=['orig', 'fail'],
        dtype=object,
    )

    df['empty1'] = None
    df['is_equal'] = df.apply(lambda x: (lambda a, b: None if a is None or b is None else (a==b)+0)(x['orig'], x['fail']), axis=1)
    df['empty2'] = None
    df['is_equal_o_f_prev_1'] = pd.Series([None if a is None or b is None else (a==b)+0 for a, b in zip(df['orig'].values[:-1].tolist()+[None], df['fail'].values[1:].tolist()+[None]*1)], dtype=object)
    df['is_equal_o_f_next_1'] = pd.Series([None if a is None or b is None else (a==b)+0 for a, b in zip(df['orig'].values[1:].tolist()+[None], df['fail'].values[:-1].tolist()+[None]*1)], dtype=object)
    df['empty3'] = None
    df['is_equal_o_f_prev_2'] = pd.Series([None if a is None or b is None else (a==b)+0 for a, b in zip(df['orig'].values[:-2].tolist()+[None], df['fail'].values[2:].tolist()+[None]*2)], dtype=object)
    df['is_equal_o_f_next_2'] = pd.Series([None if a is None or b is None else (a==b)+0 for a, b in zip(df['orig'].values[2:].tolist()+[None], df['fail'].values[:-2].tolist()+[None]*2)], dtype=object)

    # TODO: finish this tomorrow!
    arr_idx_alterning_orig = np.hstack(((0, ), np.where(arr_orig[:-1]!=arr_orig[1:])[0]+1, (arr_orig.shape[0], )))
    arr_idx_alterning_fail = np.hstack(((0, ), np.where(arr_fail[:-1]!=arr_fail[1:])[0]+1, (arr_fail.shape[0], )))

    wb = openpyxl.Workbook()
    del wb['Sheet']

    ws = wb.create_sheet('Transmission Bit Error')
    cl = ws.cell

    for column, column_name in enumerate(df.columns.values, 1):
        c = cl(column=column, row=1)
        c.value = column_name

    for (row, arr_row), (idx_orig, row_orig) in zip(enumerate(df.values, 2), df_orig.iterrows()):
        for column, val in enumerate(arr_row, 1):
            c = cl(column=column, row=row)
            c.value = val

            if val == 1:
                c.fill = PatternFill(start_color="f85050", end_color="f85050", fill_type="solid")
            elif val == 0:
                c.fill = PatternFill(start_color="538dd5", end_color="538dd5", fill_type="solid")

            if column == 2:
                if row_orig['flip_bit'] == 1:
                    c.fill = PatternFill(start_color="76933c", end_color="76933c", fill_type="solid")
                elif row_orig['insert_bit'] == 1:
                    c.fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")  
                
                if row_orig['delete_bit_prev'] == 1:
                    c.border = thin_border
                    # c.fill = PatternFill(start_color="76933c", end_color="76933c", fill_type="solid") 


    for i in range(1, df.shape[1]+1):
        ws.column_dimensions[get_column_letter(i)].width = 5.2

    wb.save(XLSX_TEMP_DIR+'transmission_bit_error.xlsx')
