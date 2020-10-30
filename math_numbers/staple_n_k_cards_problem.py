#! /usr/bin/python3.8

# -*- coding: utf-8 -*-

# Some other needed imports
import datetime
import dill
import gzip
import os
import pdb
import re
import sys
import traceback

import numpy as np
import pandas as pd
from pprint import pprint

from copy import deepcopy, copy
from dotmap import DotMap
from functools import reduce
from memory_tempfile import MemoryTempfile
from shutil import copyfile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

PATH_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))+"/"
HOME_DIR = os.path.expanduser("~")+"/"
TEMP_DIR = MemoryTempfile().gettempdir()+"/"


def mkdirs(path):
    if not os.path.exists(path):
        os.makedirs(path)


OBJS_DIR_PATH = PATH_ROOT_DIR+'objs/'
mkdirs(OBJS_DIR_PATH)

XLSX_TEMP_DIR = os.path.join(TEMP_DIR, 'xlsx_files')+'/'
mkdirs(XLSX_TEMP_DIR)

def get_cycle_length(n: int, k: int) -> int:
    # n = 3 # amount of cards per pile
    # k = 6 # amount of piles

    l = tuple([n]*k)
    l_l = [l]

    while True:
        l0 = [i-1 for i in l]
        amount_piles = len(l)

        l = tuple(sorted([i for i in l0 if i > 0] + [amount_piles]))

        if l in l_l:
            l_l.append(l)
            break

        l_l.append(l)

    cycle_length = len(l_l) - 1 - l_l.index(l_l[-1])
    print("l_l:\n")
    pprint(l_l)
    return cycle_length


if __name__ == '__main__':
    get_cycle_length(n=4, k=8)

    sys.exit()

    max_n = 30
    max_k = 30
    arr = np.zeros((max_n, max_k), dtype=object)
    n = 3 # amount of cards per pile
    k = 6 # amount of piles

    for n in range(1, max_n+1):
        print("n: {}".format(n))
        for k in range(1, max_k+1):
            cycle_length = get_cycle_length(n=n, k=k)
            arr[n-1, k-1] = cycle_length

    print("arr:\n{}".format(arr))

    wb = openpyxl.Workbook()
    del wb['Sheet']

    ws = wb.create_sheet('Cycle Length n-k piles')
    cl = ws.cell
        
    for row, arr_row in enumerate(arr, 2):
        for column, val in enumerate(arr_row, 2):
            c = cl(column=column, row=row)
            c.value = val
            
            if val == 1:
                c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    c = cl(column=1, row=1)
    c.value = 'v n / > k'

    for i in range(2, n+2):
        c = cl(column=1, row=i)
        c.value = i-1

    for i in range(2, k+2):
        c = cl(column=i, row=1)
        c.value = i-1

    for i in range(1, k+2):
        ws.column_dimensions[get_column_letter(i)].width = 5.2

    wb.save(XLSX_TEMP_DIR+'cycle_length_n_k_piles.xlsx')

    # q = 'abcdefghijklm'
    # w = 'ehgf'

    # q_1 = 'efgh'
    # w = 'ehgf' -> 'efgh'
