#! /usr/bin/env -S /usr/bin/time /usr/bin/python3.11 -i

# -*- coding: utf-8 -*-

# Some other needed imports
import datetime
import dill
import gzip
import itertools
import openpyxl
import os
import pdb
import re
import sys
import time
import traceback

import numpy as np # need installation from pip
import pandas as pd # need installation from pip
import pandasql as ps
import multiprocessing as mp

import matplotlib.pyplot as plt # need installation from pip
import pyarrow.feather as feather

from collections import defaultdict
from copy import deepcopy, copy
from dotmap import DotMap # need installation from pip
from enum import Enum, unique
from functools import reduce
from hashlib import sha256
from io import BytesIO
from memory_tempfile import MemoryTempfile # need installation from pip
from shutil import copyfile
from pprint import pprint
from typing import List, Set, Tuple, Dict, Union, Any
from PIL import Image

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

CURRENT_WORKING_DIR = os.getcwd()
PATH_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
HOME_DIR = os.path.expanduser("~")
TEMP_DIR = MemoryTempfile().gettempdir()
PYTHON_PROGRAMS_DIR = os.path.join(HOME_DIR, 'git/python_programs')

# set the relative/absolute path where the utils_load_module.py file is placed!
sys.path.append(PYTHON_PROGRAMS_DIR)
from utils_load_module import load_module_dynamically

var_glob = globals()
load_module_dynamically(**dict(var_glob=var_glob, name='utils', path=os.path.join(PYTHON_PROGRAMS_DIR, "utils.py")))
load_module_dynamically(**dict(var_glob=var_glob, name='utils_multiprocessing_manager', path=os.path.join(PYTHON_PROGRAMS_DIR, "utils_multiprocessing_manager.py")))
load_module_dynamically(**dict(var_glob=var_glob, name='utils_wb', path=os.path.join(PYTHON_PROGRAMS_DIR, "utils/utils_wb.py")))
# https://github.com/hearues-zueke-github/LCGMulXorPRNG.git
load_module_dynamically(**dict(var_glob=var_glob, name='prng', path=os.path.join(PYTHON_PROGRAMS_DIR, "../LCGMulXorPRNG/src/python/prng.py")))

mkdirs = utils.mkdirs
MultiprocessingManager = utils_multiprocessing_manager.MultiprocessingManager
RandomNumberDevice = prng.RandomNumberDevice

OBJS_DIR_PATH = os.path.join(PATH_ROOT_DIR, 'objs')
mkdirs(OBJS_DIR_PATH)

PLOTS_DIR_PATH = os.path.join(PATH_ROOT_DIR, 'plots')
mkdirs(PLOTS_DIR_PATH)

# def interpolate_prev_and_next_val(val_1, val_2):
# 	diff = abs(val_1 - val_2)
	
# 	val_left = val_1 - diff
# 	val_right = val_2 + diff

# 	return val_left, val_right


def get_l_cluster(arr: np.ndarray) -> List[Tuple[int, int, int, bool]]:
	if len(arr) == 0:
		return []
	elif len(arr) == 1:
		return [(0, 1, 0, True)]

	cluster_nr = 0

	@unique
	class State(Enum):
		IS_LE = 1
		IS_GT = 2
		SET_NEXT_VAL = 3

	l_cluster = []
	idx_prev = 0
	val_prev_1 = arr[0]
	val_prev_2 = arr[1]
	
	if val_prev_1 <= val_prev_2:
		state = State.IS_LE.value
	else:
		state = State.IS_GT.value

	idx = 2
	while idx < len(arr):
		val = arr[idx]
		if state == State.IS_LE.value:
			if val_prev_2 <= val:
				val_prev_2 = val
				idx += 1
				continue
			l_cluster.append((idx_prev, idx, cluster_nr, True))
			idx_prev = idx
			cluster_nr += 1
			val_prev_1 = val
			state = State.SET_NEXT_VAL.value
		elif state == State.IS_GT.value:
			if val_prev_2 > val:
				val_prev_2 = val
				idx += 1
				continue
			l_cluster.append((idx_prev, idx, cluster_nr, False))
			idx_prev = idx
			cluster_nr += 1
			val_prev_1 = val
			state = State.SET_NEXT_VAL.value
		elif state == State.SET_NEXT_VAL.value:
			val_prev_2 = arr[idx]
			if val_prev_1 <= val_prev_2:
				state = State.IS_LE.value
			else:
				state = State.IS_GT.value

		idx += 1

	if state == State.IS_LE.value:
		l_cluster.append((idx_prev, idx, cluster_nr, True))
	elif state == State.IS_GT.value:
		l_cluster.append((idx_prev, idx, cluster_nr, False))
	elif state == State.SET_NEXT_VAL.value:
		l_cluster.append((idx_prev, idx, cluster_nr, True))
	else:
		assert False and "should never happen!"

	return l_cluster


def sort_part_first(arr_src, arr_dst, idx_1, idx_2, iter_1, iter_2):
	idx = idx_1

	idx_val_1 = next(iter_1)
	idx_val_2 = next(iter_2)

	val_1 = arr_src[idx_val_1]
	val_2 = arr_src[idx_val_2]

	while idx < idx_2:
		if val_1 <= val_2:
			arr_dst[idx] = val_1
			idx_val_1 = next(iter_1, None)
			if idx_val_1 is None:
				idx += 1
				break
			val_1 = arr_src[idx_val_1]
		else:
			arr_dst[idx] = val_2
			idx_val_2 = next(iter_2, None)
			if idx_val_2 is None:
				idx += 1
				break
			val_2 = arr_src[idx_val_2]

		idx += 1

	if idx < idx_2:
		if idx_val_1 is None:
			arr_dst[idx] = val_2
			idx += 1
			for idx_val_2 in iter_2:
				val_2 = arr_src[idx_val_2]
				arr_dst[idx] = val_2
				idx += 1
		elif idx_val_2 is None:
			arr_dst[idx] = val_1
			idx += 1
			for idx_val_1 in iter_1:
				val_1 = arr_src[idx_val_1]
				arr_dst[idx] = val_1
				idx += 1
		else:
			assert False and "should never happen!"


def merge_sort_first_clusters(arr: np.ndarray) -> np.ndarray:
	if len(arr) < 2:
		return arr.copy()

	l_cluster_first = get_l_cluster(arr=arr)

	# for temporary sorted array
	arr_temp_1 = arr.copy()
	arr_temp_2 = np.empty(arr.shape, dtype=arr.dtype)

	# algorithm: take 2 clusters and merge them iterative

	print(f"len(l_cluster_first): {len(l_cluster_first)}")

	l_cluster_new = []
	idx_cluster = 2
	while idx_cluster <= len(l_cluster_first):
		cluster_1 = l_cluster_first[idx_cluster - 2]
		cluster_2 = l_cluster_first[idx_cluster - 1]

		# is_le is only needed for the first run!
		idx_1_1, idx_1_2, _, is_le_1 = cluster_1
		idx_2_1, idx_2_2, _, is_le_2 = cluster_2

		if is_le_1:
			iter_1 = iter(range(idx_1_1, idx_1_2, 1))
		else:
			iter_1 = iter(range(idx_1_2 - 1, idx_1_1 - 1, -1))

		if is_le_2:
			iter_2 = iter(range(idx_2_1, idx_2_2, 1))
		else:
			iter_2 = iter(range(idx_2_2 - 1, idx_2_1 - 1, -1))

		sort_part_first(
			arr_src=arr_temp_1,
			arr_dst=arr_temp_2,
			idx_1=idx_1_1,
			idx_2=idx_2_2,
			iter_1=iter_1,
			iter_2=iter_2,
		)

		l_cluster_new.append((idx_1_1, idx_2_2))
		idx_cluster += 2

	# do the last cluster separate
	if len(l_cluster_first) % 2 == 1:
		idx_1, idx_2, _, is_le = l_cluster_first[-1]
		if is_le:
			iter_idx = iter(range(idx_1, idx_2, 1))
		else:
			iter_idx = iter(range(idx_2 - 1, idx_1 - 1, -1))

		for idx_val_orig, idx_val in zip(range(idx_1, idx_2), iter_idx):
			arr_temp_2[idx_val_orig] = arr_temp_1[idx_val]

		l_cluster_new.append((idx_1, idx_2))

	arr_temp_1, arr_temp_2 = arr_temp_2, arr_temp_1

	l_cluster = l_cluster_new

	# print(f"len(l_cluster): {len(l_cluster)}")
	while len(l_cluster) > 1:
		l_cluster_new =[]
		idx_cluster = 2
		while idx_cluster <= len(l_cluster):
			idx_1_1, idx_1_2 = l_cluster[idx_cluster - 2]
			idx_2_1, idx_2_2 = l_cluster[idx_cluster - 1]

			iter_1 = iter(range(idx_1_1, idx_1_2, 1))
			iter_2 = iter(range(idx_2_1, idx_2_2, 1))

			sort_part_first(
				arr_src=arr_temp_1,
				arr_dst=arr_temp_2,
				idx_1=idx_1_1,
				idx_2=idx_2_2,
				iter_1=iter_1,
				iter_2=iter_2,
			)

			l_cluster_new.append((idx_1_1, idx_2_2))
			idx_cluster += 2

			if idx_2_2 - idx_1_1 > 1:
				assert np.all(np.diff(arr_temp_2[idx_1_1:idx_2_2]) >= 0)

		# do the last cluster separate
		if len(l_cluster) % 2 == 1:
			idx_1, idx_2 = l_cluster[-1]

			for idx_val in range(idx_1, idx_2, 1):
				arr_temp_2[idx_val] = arr_temp_1[idx_val]
			l_cluster_new.append((idx_1, idx_2))

		l_cluster = l_cluster_new
		arr_temp_1, arr_temp_2 = arr_temp_2, arr_temp_1
		# print(f"len(l_cluster): {len(l_cluster)}")

	return arr_temp_1


if __name__ == '__main__':
	# for amount in range(7, 2000):
	amount = 1000
	print(f"amount: {amount}")
	seed_u8 = np.array([0x00, 0x02], dtype=np.uint8)
	length_u8 = 128

	rnd = RandomNumberDevice(
		seed_u8=seed_u8,
		length_u8=length_u8,
	)

	arr_u64 = rnd.calc_next_uint64(amount=amount)
	# print(f"arr_u64: {arr_u64}")

	arr_i16 = arr_u64.view(np.int16)
	print(f"arr_i16: {arr_i16}")

	arr_raw = arr_i16.astype(np.int32)
	
	# val_first, _ = interpolate_prev_and_next_val(arr_raw[0], arr_raw[1])
	# _, val_last = interpolate_prev_and_next_val(arr_raw[-2], arr_raw[-1])

	arr = arr_raw
	# arr = np.hstack((
	# 	# (val_first, ),
	# 	arr_raw,
	# 	# (val_last, ),
	# ))
	
	arr_sort = merge_sort_first_clusters(arr=arr)
	assert np.all(np.sort(arr) == arr_sort)

	df = pd.DataFrame(dtype=object)
	df['arr'] = arr

	# add this part into a separate function maybe...
	arr = df['arr'].values
	arr_lt = np.hstack(((arr[:-1] < arr[1:]) + 0, (1, )))

	df['arr_lt'] = arr_lt

	wb = openpyxl.Workbook()
	del wb['Sheet']

	list_build = [df.columns.tolist()] + df.values.tolist()
	utils_wb.create_new_sheet(wb=wb, sheet_name='arr_generate', list_build=list_build, column_widths=[8.0 for _ in df.columns])

	ws = wb['arr_generate']

	green_fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
	red_fill = PatternFill(start_color='da9694', end_color='da9694', fill_type='solid')
	
	ws.conditional_formatting.add('B2:B161', CellIsRule(operator='equal', formula=['0'], fill=red_fill))
	ws.conditional_formatting.add('B2:B161', CellIsRule(operator='equal', formula=['1'], fill=green_fill))

	l_cluster = get_l_cluster(arr=arr)
	l_data = [(cluster_nr, int(is_le), ','.join(map(str, [arr[i] for i in range(idx_1, idx_2)]))) for idx_1, idx_2, cluster_nr, is_le in l_cluster]
	df_cluster = pd.DataFrame(data=l_data, columns=['cluster_nr', 'is_le', 'numbers'], dtype=object)
	list_build = [df_cluster.columns.tolist()] + df_cluster.values.tolist()
	utils_wb.create_new_sheet(wb=wb, sheet_name='clusters', list_build=list_build, column_widths=[8.0 for _ in df_cluster.columns])

	xlsx_tmp_file_path = os.path.join(TEMP_DIR, 'arr_example.xlsx')
	print(f"xlsx_tmp_file_path: {xlsx_tmp_file_path}")

	wb.save(xlsx_tmp_file_path)
